

from  openpyxl.reader.excel import load_workbook

def readxls(path):
    dic={}
    file=load_workbook(filename=path)
    # print(file.sheetnames)
    sheets=file.sheetnames

    for sheetName in  sheets:
        sheet=file[sheetName]
        # 一张表所有数据
        sheetInfo=[]
        # # 最大行数
        # print(sheet.max_row)
        # # 最大列数
        # print(sheet.max_column)
        # # 表名
        # print(sheet.title)
        for linNum in range(1, sheet.max_row + 1):
            # print(linNum)
            lineList = []
            for columnNum in range(1, sheet.max_column + 1):
                value = sheet.cell(row=linNum, column=columnNum).value
                lineList.append(value)
            # print(lineList)
            sheetInfo.append(lineList)
        dic[sheetName]=sheetInfo
    return dic
path="/Users/zzs/Desktop/SVN/trunk/functions/客户资料/移动OA功能链接.xlsx"
print(readxls(path))